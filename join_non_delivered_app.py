import streamlit as st
import base64
from io import BytesIO
import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import PatternFill, NamedStyle, Font, Border, Side
import random

# Load contacts function (securely retrieves the secret file)
@st.cache_data
def load_contacts():
    encoded_data = st.secrets["all_contacts"]["content"]
    decoded_bytes = base64.b64decode(encoded_data)
    return pd.read_excel(BytesIO(decoded_bytes), engine='openpyxl')

# Function to load uploaded file in multiple formats
def load_uploaded_file(uploaded_file):
    if uploaded_file.name.endswith('.csv'):
        return pd.read_csv(uploaded_file)
    elif uploaded_file.name.endswith(('.xlsx', '.xls')):
        return pd.read_excel(uploaded_file, engine='openpyxl')
    else:
        st.error("Unsupported file format. Please upload a .csv, .xls or .xlsx file.")
        st.stop()

# Set background image function
def set_background(image_path):
    with open(image_path, "rb") as f:
        encoded_image = base64.b64encode(f.read()).decode()
    
    bg_css = f"""
    <style>
    .stApp {{
        background-image: url("data:image/png;base64,{encoded_image}");
        background-size: cover;
    }}
    </style>
    """
    st.markdown(bg_css, unsafe_allow_html=True)

# Call the background function
background_image_path = "assets/background.png"
set_background(background_image_path)

# App Title
st.title("Join Non-Delivered Emails with Account & CS Owner Details")

# File upload for Non-Deliverable List
st.subheader("Upload the Non-Deliverable List (Excel or CSV)")
non_deliverable_file = st.file_uploader("Upload your file", type=["csv", "xlsx", "xls"], key="file1")

if non_deliverable_file:
    # Load the uploaded file
    non_deliverable_df = load_uploaded_file(non_deliverable_file)

    # Load the "All Contacts" file from storage
    all_contacts_df = load_contacts()

    # Perform the join operation (Left Join)
    joined_df = non_deliverable_df.merge(
        all_contacts_df,
        left_on="Recipient",
        right_on="Email",
        how="left"
    )

    # Drop the Email column since we already have Recipient
    joined_df = joined_df.drop(columns=["Email"], errors="ignore")

    # Sort by 'CS Owner' column (if exists)
    if 'CS Owner' in joined_df.columns:
        joined_df = joined_df.sort_values(by='CS Owner')

    # Display success message and preview
    st.success("Files joined successfully!")
    st.write("Preview of the joined file:")
    st.dataframe(joined_df.head(10))  # Show only 10 rows for privacy

    # ------------------------------------
    # CSV File Creation (with Table Format)
    # ------------------------------------
    csv_buffer = joined_df.to_csv(index=False).encode('utf-8')

    # ------------------------------------
    # XLSX File Creation (with Table Format and Row Coloring)
    # ------------------------------------
    excel_buffer = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Joined Data"

    # Write headers with bold format
    header_style = NamedStyle(name="header_style")
    header_style.font = Font(bold=True, color="FFFFFF")
    header_style.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
    workbook.add_named_style(header_style)

    # Write headers
    for col_num, column_title in enumerate(joined_df.columns, start=1):
        cell = sheet.cell(row=1, column=col_num, value=column_title)
        cell.style = "header_style"

    # Assign colors based on 'CS Owner'
    color_map = {}
    color_palette = [
        "FFDDDD", "DDEEFF", "E0F7FA", "FCE4EC", "FFF3E0", "E8F5E9", "F3E5F5", "E3F2FD"
    ]

    # Assign colors for unique CS Owners
    for index, row in joined_df.iterrows():
        cs_owner = row.get('CS Owner')
        if cs_owner not in color_map:
            color_map[cs_owner] = random.choice(color_palette)

    # Write data rows with colors
    for row_idx, row in enumerate(dataframe_to_rows(joined_df, index=False, header=False), start=2):
        for col_idx, value in enumerate(row, start=1):
            cell = sheet.cell(row=row_idx, column=col_idx, value=value)
            cs_owner = row[joined_df.columns.get_loc('CS Owner')] if 'CS Owner' in joined_df.columns else None
            if cs_owner in color_map:
                cell.fill = PatternFill(start_color=color_map[cs_owner],
                                        end_color=color_map[cs_owner],
                                        fill_type="solid")

    workbook.save(excel_buffer)
    excel_buffer.seek(0)

    # Provide download buttons
    st.subheader("Download Options")
    st.download_button(
        label="📥 Download as CSV - No colored rows by CS Owner ⚫⚪",
        data=csv_buffer,
        file_name="joined_file.csv",
        mime="text/csv"
    )

    st.download_button(
        label="📥 Download as Excel (XLSX) - Colored rows by CS Owner 🎨",
        data=excel_buffer,
        file_name="joined_file.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
